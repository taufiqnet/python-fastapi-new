import uuid
from datetime import date, datetime, time, timedelta

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.modules.hr_payroll.attendance.models import (
    Attendance,
    AttendanceSourceEnum,
    AttendanceStatusEnum,
)
from app.modules.hr_payroll.attendance.repository import AttendanceRepository
from app.modules.hr_payroll.attendance.schemas import (
    AttendanceCreate,
    AttendanceUpdate,
)
from io import BytesIO
import openpyxl

from app.modules.hr_payroll.employees.repository import EmployeeRepository


class AttendanceService:
    def __init__(
        self,
        repository: AttendanceRepository | None = None,
        employee_repository: EmployeeRepository | None = None,
    ):
        self.repository = repository or AttendanceRepository()
        self.employee_repository = employee_repository or EmployeeRepository()

    def _compute_hours(
        self,
        data: AttendanceCreate | AttendanceUpdate,
        existing: Attendance | None = None,
    ) -> tuple[float, float]:
        check_in = (
            data.check_in
            if data.check_in is not None
            else (existing.check_in if existing else None)
        )
        check_out = (
            data.check_out
            if data.check_out is not None
            else (existing.check_out if existing else None)
        )

        if data.work_hours is not None and data.work_hours > 0.0:
            work_hours = float(data.work_hours)
        elif check_in and check_out:
            today = date.today()
            dt_in = datetime.combine(today, check_in)
            dt_out = datetime.combine(today, check_out)
            if dt_out < dt_in:
                dt_out += timedelta(days=1)
            diff = dt_out - dt_in
            work_hours = round(diff.total_seconds() / 3600.0, 2)
        else:
            work_hours = (
                float(existing.work_hours)
                if (existing and existing.work_hours is not None)
                else 0.0
            )

        if data.overtime_hours is not None:
            overtime_hours = float(data.overtime_hours)
        elif work_hours > 8.0:
            overtime_hours = round(work_hours - 8.0, 2)
        else:
            overtime_hours = 0.0

        return work_hours, overtime_hours

    def get_records(
        self,
        db: Session,
        skip: int = 0,
        limit: int = 100,
        business_id: int | None = None,
        employee_id: uuid.UUID | None = None,
        att_date: date | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        status_filter: str | AttendanceStatusEnum | None = None,
    ) -> list[Attendance]:
        return self.repository.get_all(
            db,
            skip=skip,
            limit=limit,
            business_id=business_id,
            employee_id=employee_id,
            att_date=att_date,
            start_date=start_date,
            end_date=end_date,
            status=status_filter,
        )

    def get_record(self, db: Session, attendance_uuid: uuid.UUID) -> Attendance:
        record = self.repository.get_by_id(db, attendance_uuid)
        if not record:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Attendance record not found",
            )
        return record

    def create_record(self, db: Session, data: AttendanceCreate) -> Attendance:
        employee = self.employee_repository.get_by_id(db, data.employee_id)
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee with id '{data.employee_id}' not found",
            )

        existing = self.repository.get_by_emp_date(
            db,
            employee_id=data.employee_id,
            att_date=data.date,
            business_id=data.business_id,
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Attendance record already exists for employee on {data.date}",
            )

        work_hours, overtime_hours = self._compute_hours(data)
        data.work_hours = work_hours
        data.overtime_hours = overtime_hours

        return self.repository.create(db, data)

    def update_record(
        self,
        db: Session,
        attendance_uuid: uuid.UUID,
        data: AttendanceUpdate,
    ) -> Attendance:
        record = self.get_record(db, attendance_uuid)

        target_emp_id = data.employee_id or record.employee_id
        target_date = data.date or record.date
        target_biz_id = (
            data.business_id
            if data.business_id is not None
            else record.business_id
        )

        if (target_emp_id != record.employee_id) or (target_date != record.date):
            existing = self.repository.get_by_emp_date(
                db,
                employee_id=target_emp_id,
                att_date=target_date,
                business_id=target_biz_id,
            )
            if existing and existing.id != attendance_uuid:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        "Attendance record already exists for "
                        f"employee on {target_date}"
                    ),
                )

        work_hours, overtime_hours = self._compute_hours(data, existing=record)
        data.work_hours = work_hours
        data.overtime_hours = overtime_hours

        return self.repository.update(db, record, data)

    def check_in(
        self,
        db: Session,
        business_id: int,
        employee_id: uuid.UUID,
        att_date: date | None = None,
        check_in_time: time | None = None,
        note: str | None = None,
    ) -> Attendance:
        employee = self.employee_repository.get_by_id(db, employee_id)
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee with id '{employee_id}' not found",
            )

        now = datetime.now()
        target_date = att_date or now.date()
        target_check_in = check_in_time or now.time().replace(microsecond=0)

        existing = self.repository.get_by_emp_date(
            db,
            employee_id=employee_id,
            att_date=target_date,
            business_id=business_id,
        )

        if existing:
            update_data = AttendanceUpdate(
                check_in=target_check_in,
                note=note or existing.note,
            )
            return self.update_record(db, existing.id, update_data)

        status_val = AttendanceStatusEnum.PRESENT
        if target_check_in > time(9, 15):
            status_val = AttendanceStatusEnum.LATE

        create_data = AttendanceCreate(
            business_id=business_id,
            employee_id=employee_id,
            date=target_date,
            status=status_val,
            check_in=target_check_in,
            source=AttendanceSourceEnum.SYSTEM,
            note=note,
        )
        return self.create_record(db, create_data)

    def check_out(
        self,
        db: Session,
        business_id: int,
        employee_id: uuid.UUID,
        att_date: date | None = None,
        check_out_time: time | None = None,
        note: str | None = None,
    ) -> Attendance:
        employee = self.employee_repository.get_by_id(db, employee_id)
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee with id '{employee_id}' not found",
            )

        now = datetime.now()
        target_date = att_date or now.date()
        target_check_out = check_out_time or now.time().replace(microsecond=0)

        existing = self.repository.get_by_emp_date(
            db,
            employee_id=employee_id,
            att_date=target_date,
            business_id=business_id,
        )

        if not existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"No check-in record found for employee on {target_date}. "
                    "Please check in first."
                ),
            )

        update_data = AttendanceUpdate(
            check_out=target_check_out,
            note=note or existing.note,
        )
        return self.update_record(db, existing.id, update_data)

    def delete_record(self, db: Session, attendance_uuid: uuid.UUID) -> None:
        record = self.get_record(db, attendance_uuid)
        self.repository.delete(db, record)

    def generate_excel_template(self, db: Session, business_id: int) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Template"

        headers = [
            "Employee ID",
            "Employee Name",
            "Date (YYYY-MM-DD)",
            "Status (present/absent/late/half_day/on_leave/holiday/weekend)",
            "Check In (HH:MM)",
            "Check Out (HH:MM)",
            "Work Hours",
            "Overtime Hours",
            "Note",
        ]
        ws.append(headers)

        employees = self.employee_repository.get_all(db, business_id=business_id, limit=500)
        sample_date = date.today().strftime("%Y-%m-%d")

        if employees:
            for emp in employees[:5]:
                ws.append([
                    emp.employee_id,
                    emp.full_name,
                    sample_date,
                    "present",
                    "09:00",
                    "18:00",
                    8.0,
                    1.0,
                    "Bulk imported attendance",
                ])
        else:
            ws.append([
                "EMP001",
                "John Doe",
                sample_date,
                "present",
                "09:00",
                "18:00",
                8.0,
                1.0,
                "Sample entry",
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def import_attendance_excel(
        self, db: Session, business_id: int, file_bytes: bytes
    ) -> dict[str, int | list[str]]:
        try:
            wb = openpyxl.load_workbook(filename=BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file format: {str(e)}",
            )

        employees = self.employee_repository.get_all(db, business_id=business_id, limit=2000)
        emp_map = {str(e.employee_id).strip().lower(): e for e in employees}

        success_count = 0
        error_messages: list[str] = []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file is empty.",
            )

        # Header detection
        header = [str(cell or "").strip().lower() for cell in rows[0]]
        start_idx = 1 if "employee id" in header or "employee_id" in header or "date" in header else 0

        for row_idx, row in enumerate(rows[start_idx:], start=start_idx + 1):
            if not row or not any(row):
                continue

            emp_code_raw = str(row[0] or "").strip()
            date_raw = row[2] if len(row) > 2 else None
            status_raw = str(row[3] or "present").strip().lower() if len(row) > 3 and row[3] else "present"
            check_in_raw = row[4] if len(row) > 4 else None
            check_out_raw = row[5] if len(row) > 5 else None
            work_hours_raw = row[6] if len(row) > 6 else None
            overtime_hours_raw = row[7] if len(row) > 7 else None
            note_raw = str(row[8] or "").strip() if len(row) > 8 and row[8] else "Imported via Excel"

            if not emp_code_raw:
                error_messages.append(f"Row {row_idx}: Missing Employee ID.")
                continue

            emp = emp_map.get(emp_code_raw.lower())
            if not emp:
                error_messages.append(f"Row {row_idx}: Employee with ID '{emp_code_raw}' not found.")
                continue

            # Parse date
            att_date: date | None = None
            if isinstance(date_raw, (datetime, date)):
                att_date = date_raw.date() if isinstance(date_raw, datetime) else date_raw
            elif isinstance(date_raw, str) and date_raw.strip():
                try:
                    att_date = datetime.strptime(date_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    try:
                        att_date = datetime.strptime(date_raw.strip(), "%m/%d/%Y").date()
                    except ValueError:
                        pass

            if not att_date:
                error_messages.append(f"Row {row_idx}: Invalid date format '{date_raw}'. Expected YYYY-MM-DD.")
                continue

            # Parse check-in / check-out times
            def parse_time_val(val) -> time | None:
                if isinstance(val, time):
                    return val
                if isinstance(val, datetime):
                    return val.time()
                if isinstance(val, str) and val.strip():
                    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
                        try:
                            return datetime.strptime(val.strip(), fmt).time()
                        except ValueError:
                            pass
                return None

            check_in_time = parse_time_val(check_in_raw)
            check_out_time = parse_time_val(check_out_raw)

            # Map status
            try:
                status_enum = AttendanceStatusEnum(status_raw)
            except ValueError:
                status_enum = AttendanceStatusEnum.PRESENT

            # Work / Overtime hours
            try:
                work_hrs = float(work_hours_raw) if work_hours_raw is not None and str(work_hours_raw).strip() != "" else None
            except (ValueError, TypeError):
                work_hrs = None

            try:
                ot_hrs = float(overtime_hours_raw) if overtime_hours_raw is not None and str(overtime_hours_raw).strip() != "" else None
            except (ValueError, TypeError):
                ot_hrs = None

            existing = self.repository.get_by_emp_date(
                db, employee_id=emp.id, att_date=att_date, business_id=business_id
            )

            if existing:
                upd = AttendanceUpdate(
                    status=status_enum,
                    check_in=check_in_time,
                    check_out=check_out_time,
                    work_hours=work_hrs,
                    overtime_hours=ot_hrs,
                    note=note_raw,
                )
                self.update_record(db, existing.id, upd)
                success_count += 1
            else:
                crt = AttendanceCreate(
                    business_id=business_id,
                    employee_id=emp.id,
                    date=att_date,
                    status=status_enum,
                    check_in=check_in_time,
                    check_out=check_out_time,
                    work_hours=work_hrs or 0.0,
                    overtime_hours=ot_hrs or 0.0,
                    source=AttendanceSourceEnum.MANUAL,
                    note=note_raw,
                )
                self.create_record(db, crt)
                success_count += 1

        return {
            "imported_count": success_count,
            "errors": error_messages,
        }
